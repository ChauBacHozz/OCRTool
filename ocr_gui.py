import threading
from pathlib import Path

import customtkinter as ctk
from tkinter import filedialog, messagebox

from qwen_ocr_viet import ocr_vietnamese

ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")


class OCRApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("OCR Tool - Qwen3.5")
        self.geometry("780x560")

        self.files: list[str] = []
        self.results: list[dict] = []

        # ── File list ──
        top = ctk.CTkFrame(self)
        top.pack(fill="x", padx=12, pady=(12, 0))

        ctk.CTkLabel(top, text="Images:", anchor="w", font=("", 13)).pack(side="left")
        ctk.CTkButton(top, text="+ Add", width=70, command=self.add_files).pack(side="right", padx=(2, 0))
        ctk.CTkButton(top, text="Clear", width=70, command=self.clear_files).pack(side="right", padx=(2, 0))

        self.listbox = ctk.CTkTextbox(self, height=180, font=("", 12))
        self.listbox.pack(fill="both", expand=True, padx=12, pady=6)
        self.listbox.configure(state="disabled")

        # ── Output dir ──
        dir_frame = ctk.CTkFrame(self)
        dir_frame.pack(fill="x", padx=12, pady=(0, 6))

        ctk.CTkLabel(dir_frame, text="Save to:", font=("", 13)).pack(side="left")
        self.dir_label = ctk.CTkLabel(dir_frame, text="(same folder as images)", text_color="gray")
        self.dir_label.pack(side="left", padx=6)
        ctk.CTkButton(dir_frame, text="Browse", width=80, command=self.choose_output_dir).pack(side="right")
        self.output_dir = ""

        # ── Progress ──
        self.progress = ctk.CTkProgressBar(self)
        self.progress.pack(fill="x", padx=12, pady=4)
        self.progress.set(0)

        self.status_var = ctk.StringVar(value="Ready")
        ctk.CTkLabel(self, textvariable=self.status_var, anchor="w", text_color="gray").pack(fill="x", padx=12)

        # ── Log ──
        self.log_box = ctk.CTkTextbox(self, height=120, font=("", 11))
        self.log_box.pack(fill="both", padx=12, pady=6)
        self.log_box.configure(state="disabled")

        # ── Run ──
        self.run_btn = ctk.CTkButton(self, text="Run OCR", height=38, command=self.run_ocr)
        self.run_btn.pack(fill="x", padx=12, pady=(0, 12))

    # ──────────────────────────────────────────────

    def log(self, msg: str):
        self.log_box.configure(state="normal")
        self.log_box.insert("end", msg + "\n")
        self.log_box.see("end")
        self.log_box.configure(state="disabled")

    def add_files(self):
        paths = filedialog.askopenfilenames(
            title="Select images",
            filetypes=[("Images", "*.png *.jpg *.jpeg *.bmp *.tiff *.webp")],
        )
        if not paths:
            return
        self.listbox.configure(state="normal")
        for p in paths:
            if p not in self.files:
                self.files.append(p)
                self.listbox.insert("end", Path(p).name + "\n")
        self.listbox.configure(state="disabled")
        self.progress.set(0)

    def clear_files(self):
        self.files.clear()
        self.listbox.configure(state="normal")
        self.listbox.delete("1.0", "end")
        self.listbox.configure(state="disabled")
        self.progress.set(0)
        self.status_var.set("Ready")

    def choose_output_dir(self):
        d = filedialog.askdirectory(title="Select output folder")
        if d:
            self.output_dir = d
            self.dir_label.configure(text=d, text_color="white")

    # ──────────────────────────────────────────────

    def run_ocr(self):
        if not self.files:
            messagebox.showwarning("No files", "Please add at least one image.")
            return
        self.run_btn.configure(state="disabled")
        self.progress.set(0)
        self.log_box.configure(state="normal")
        self.log_box.delete("1.0", "end")
        self.log_box.configure(state="disabled")
        self.results.clear()

        def task():
            n = len(self.files)
            for i, f in enumerate(self.files):
                name = Path(f).name
                self.after(0, self.log, f"[{i+1}/{n}] {name} ...")
                self.after(0, self.status_var.set, f"[{i+1}/{n}] Processing: {name}")

                try:
                    data = ocr_vietnamese(f)
                    self.results.append({"file": name, **data.model_dump()})
                    self.after(0, self.log, f"       -> OK")
                except Exception as e:
                    self.after(0, self.log, f"       -> ERROR: {e}")

                self.after(0, lambda v=(i+1)/n: self.progress.set(v))

            self.after(0, self._save_xlsx)
            self.after(0, self._done)

        threading.Thread(target=task, daemon=True).start()

    def _save_xlsx(self):
        if not self.results:
            self.log("No results to save.")
            return

        import openpyxl
        from openpyxl.utils import get_column_letter

        out_dir = self.output_dir or Path(self.files[0]).parent
        out_path = Path(out_dir) / "ocr_results.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "OCR Results"

        headers = list(self.results[0].keys())
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)

        for r, row_data in enumerate(self.results, 2):
            for c, key in enumerate(headers, 1):
                ws.cell(row=r, column=c, value=row_data.get(key, ""))

        for col in range(1, len(headers) + 1):
            max_len = max(
                len(str(ws.cell(row=r, column=col).value or ""))
                for r in range(1, len(self.results) + 2)
            )
            ws.column_dimensions[get_column_letter(col)].width = min(max_len + 4, 60)

        wb.save(out_path)
        self.log(f"Saved: {out_path}")

    def _done(self):
        self.run_btn.configure(state="normal")
        self.status_var.set("Done.")
        self.log("=== All done ===")
        messagebox.showinfo("Complete", "All files processed.")


if __name__ == "__main__":
    OCRApp().mainloop()
